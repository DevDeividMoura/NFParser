import pytest
import os
import shutil
import openpyxl
from pathlib import Path
from unittest.mock import Mock, patch
from src.data_exporter import save_to_excel

@pytest.fixture
def tmp_path():
    path = Path(__file__).resolve().parent / "tmp"
    path.mkdir(exist_ok=True)  # Cria o diretório se não existir
    yield path  # Fornece o diretório para o teste
    shutil.rmtree(path)

@pytest.fixture
def sample_data():
    return [
        {"date": "30/11/2024", "amount": "425,01", "plate": "ABC-1234", "km": "62876"}
    ]

@pytest.fixture
def expected_headers():
    return ["date", "station", "amount", "plate", "km"]

@pytest.fixture
def expected_station():
    return "deluca"

### Teste para a função save_to_excel ###
def test_save_to_excel_creates_file_correctly(sample_data, expected_headers, expected_station, tmp_path):
    file_path = tmp_path / "output.xlsx"
    save_to_excel(sample_data, str(file_path))
    
    # Verificar se o arquivo foi gerado
    assert os.path.exists(file_path), "Excel file was not created"

    # Abrir o arquivo Excel para verificar o conteúdo
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Verificar cabeçalhos
    headers = [cell.value for cell in sheet[1]]
    assert headers == expected_headers, "Headers do not match"

    # Verificar dados
    first_row = [cell.value for cell in sheet[2]]
    expected_row = [
        sample_data[0]["date"], expected_station, sample_data[0]["amount"], 
        sample_data[0]["plate"], sample_data[0]["km"]
    ]
    assert first_row == expected_row, "Row data does not match"

def test_save_to_excel_with_empty_data(tmp_path):
    file_path = tmp_path / "output.xlsx"
    save_to_excel([], str(file_path))
    
    # Verificar se o arquivo foi gerado
    assert os.path.exists(file_path), "Excel file was not created"

    # Abrir o arquivo Excel para verificar o conteúdo
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Verificar que não há dados
    assert sheet.max_row == 1, "Sheet should only contain headers"
    assert sheet.max_column == 5, "Sheet should not contain any headers"    

def test_save_to_excel_with_multiple_rows(sample_data, expected_headers, expected_station, tmp_path):
    additional_data = [
        {"date": "01/12/2024", "amount": "300,00", "plate": "XYZ-5678", "km": "70000"},
        {"date": "02/12/2024", "amount": "150,50", "plate": "N/A", "km": "N/A"}
    ]
    sample_data.extend(additional_data)
    file_path = tmp_path / "output.xlsx"
    save_to_excel(sample_data, str(file_path))
    
    # Verificar se o arquivo foi gerado
    assert os.path.exists(file_path), "Excel file was not created"

    # Abrir o arquivo Excel para verificar o conteúdo
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Verificar cabeçalhos
    headers = [cell.value for cell in sheet[1]]
    assert headers == expected_headers, "Headers do not match"

    # Verificar dados
    for i, data in enumerate(sample_data, start=2):
        row = [cell.value for cell in sheet[i]]
        expected_row = [
            data["date"], expected_station, data["amount"], 
            data["plate"], data["km"]
        ]
        assert row == expected_row, f"Row {i} data does not match"

    # Verificar cabeçalhos
    headers = [cell.value for cell in sheet[1]]
    assert headers == expected_headers, "Headers do not match"

    # Verificar dados
    first_row = [cell.value for cell in sheet[2]]
    expected_row = [
        sample_data[0]["date"], expected_station, sample_data[0]["amount"], 
        sample_data[0]["plate"], sample_data[0]["km"]
    ]
    assert first_row == expected_row, "Row data does not match"

